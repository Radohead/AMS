"""
Excel/CSV 批量导出工具
支持导出资产、员工等数据
"""
from typing import Optional, List, Dict, Any
from datetime import datetime
from io import BytesIO, StringIO
import csv

from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.models.asset import Asset, Category, Department, Employee


# 导出字段配置
ASSET_EXPORT_FIELDS = [
    ("资产编码", "asset_no"),
    ("资产名称", "name"),
    ("分类", "category_name"),
    ("资产类型", "asset_type"),
    ("状态", "status"),
    ("品牌", "brand"),
    ("型号", "model"),
    ("序列号", "serial_no"),
    ("购买日期", "purchase_date"),
    ("购买价格", "purchase_price"),
    ("保修截止", "warranty_end"),
    ("使用部门", "department_name"),
    ("使用人", "user_name"),
    ("保管人", "keeper_name"),
    ("存放位置", "location"),
    ("数量", "quantity"),
    ("单位", "unit"),
    ("描述", "description"),
    ("备注", "remarks"),
    # 房地产专用字段
    ("详细地址", "address"),
    ("建筑面积", "area"),
    ("占地面积", "land_area"),
    ("产权类型", "property_type"),
    ("产权证号", "property_no"),
    ("土地证号", "land_no"),
    ("楼栋号", "building_no"),
    ("楼层", "floor"),
    ("房间号", "room_no"),
    ("用途", "usage"),
    ("建成年份", "build_year"),
    ("建筑结构", "structure"),
    ("创建时间", "created_at"),
]

# 资产类型中文映射
ASSET_TYPE_NAMES = {
    "fixed": "固定资产",
    "consumable": "易耗品",
    "real_estate": "房地产",
}

# 资产状态中文映射
ASSET_STATUS_NAMES = {
    "stock": "在库",
    "in_use": "使用中",
    "repair": "维修中",
    "scrapped": "已报废",
    "lost": "丢失",
}


def get_field_value(asset: Asset, field_name: str) -> Any:
    """获取资产字段值"""
    if field_name == "category_name":
        return asset.category.name if asset.category else ""
    elif field_name == "department_name":
        return asset.department.name if asset.department else ""
    elif field_name == "user_name":
        return asset.user.name if asset.user else ""
    elif field_name == "keeper_name":
        return asset.keeper.name if asset.keeper else ""
    elif field_name == "asset_type":
        return ASSET_TYPE_NAMES.get(asset.asset_type.value, asset.asset_type.value)
    elif field_name == "status":
        return ASSET_STATUS_NAMES.get(asset.status.value, asset.status.value)
    elif field_name == "purchase_date":
        return asset.purchase_date.strftime("%Y-%m-%d") if asset.purchase_date else ""
    elif field_name == "warranty_end":
        return asset.warranty_end.strftime("%Y-%m-%d") if asset.warranty_end else ""
    elif field_name == "created_at":
        return asset.created_at.strftime("%Y-%m-%d %H:%M:%S") if asset.created_at else ""
    elif hasattr(asset, field_name):
        value = getattr(asset, field_name)
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return value if value is not None else ""
    return ""


def query_assets_for_export(
    db: Session,
    keyword: Optional[str] = None,
    category_id: Optional[int] = None,
    department_id: Optional[int] = None,
    user_id: Optional[int] = None,
    status: Optional[str] = None,
    asset_type: Optional[str] = None,
) -> List[Asset]:
    """查询要导出的资产"""
    query = db.query(Asset)

    # 添加筛选条件
    if keyword:
        query = query.filter(
            Asset.name.contains(keyword) |
            Asset.asset_no.contains(keyword) |
            Asset.serial_no.contains(keyword)
        )
    if category_id:
        query = query.filter(Asset.category_id == category_id)
    if department_id:
        query = query.filter(Asset.department_id == department_id)
    if user_id:
        query = query.filter(Asset.user_id == user_id)
    if status:
        query = query.filter(Asset.status == status)
    if asset_type:
        query = query.filter(Asset.asset_type == asset_type)

    return query.order_by(Asset.created_at.desc()).all()


def export_assets_to_excel(
    db: Session,
    keyword: Optional[str] = None,
    category_id: Optional[int] = None,
    department_id: Optional[int] = None,
    user_id: Optional[int] = None,
    status: Optional[str] = None,
    asset_type: Optional[str] = None,
) -> bytes:
    """
    导出资产到 Excel 文件

    Returns:
        Excel 文件字节数据
    """
    # 查询资产
    assets = query_assets_for_export(
        db, keyword, category_id, department_id, user_id, status, asset_type
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "资产导出"

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写入表头
    headers = [h for h, _ in ASSET_EXPORT_FIELDS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 设置列宽
    col_widths = [15, 20, 12, 10, 10, 12, 15, 15, 12, 12, 12, 12, 10, 10, 15, 8, 8,
                  20, 20, 25, 12, 12, 10, 15, 15, 8, 8, 8, 8, 8, 8, 20]
    for idx, width in enumerate(col_widths[:len(headers)], start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    # 写入数据
    for row_idx, asset in enumerate(assets, start=2):
        for col_idx, (_, field_name) in enumerate(ASSET_EXPORT_FIELDS, start=1):
            value = get_field_value(asset, field_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def export_assets_to_csv(
    db: Session,
    keyword: Optional[str] = None,
    category_id: Optional[int] = None,
    department_id: Optional[int] = None,
    user_id: Optional[int] = None,
    status: Optional[str] = None,
    asset_type: Optional[str] = None,
) -> bytes:
    """
    导出资产到 CSV 文件

    Returns:
        CSV 文件字节数据
    """
    # 查询资产
    assets = query_assets_for_export(
        db, keyword, category_id, department_id, user_id, status, asset_type
    )

    # 使用 StringIO 写入文本
    text_buffer = StringIO()
    headers = [h for h, _ in ASSET_EXPORT_FIELDS]

    writer = csv.writer(text_buffer)

    # 写入 BOM（Excel 需要）- 先在开头加上
    csv_content = "\ufeff"

    # 写入表头
    writer.writerow(headers)

    # 写入数据
    for asset in assets:
        row_data = []
        for _, field_name in ASSET_EXPORT_FIELDS:
            value = get_field_value(asset, field_name)
            row_data.append(value)
        writer.writerow(row_data)

    csv_content += text_buffer.getvalue()
    return csv_content.encode("utf-8-sig")


def generate_export_filename(format: str = "xlsx") -> str:
    """生成导出文件名"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"资产导出_{timestamp}.{format}"
