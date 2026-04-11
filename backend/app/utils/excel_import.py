"""
Excel 批量导入工具
支持解析 Excel 文件并转换为资产/员工数据
"""
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session
from app.models.asset import Category, Department, Employee, AssetType, AssetStatus
from app.models.user import User


# Excel 模板表头配置
ASSET_IMPORT_HEADERS = [
    ("资产名称", "name"),
    ("分类编码", "category_code"),
    ("资产类型", "asset_type"),
    ("品牌", "brand"),
    ("型号", "model"),
    ("序列号", "serial_no"),
    ("购买日期", "purchase_date"),
    ("购买价格", "purchase_price"),
    ("保修截止", "warranty_end"),
    ("部门编码", "department_code"),
    ("存放位置", "location"),
    ("数量", "quantity"),
    ("单位", "unit"),
    ("描述", "description"),
    ("备注", "remarks"),
]

# 资产类型映射
ASSET_TYPE_MAP = {
    "固定资产": AssetType.FIXED,
    "fixed": AssetType.FIXED,
    "易耗品": AssetType.CONSUMABLE,
    "consumable": AssetType.CONSUMABLE,
    "房地产": AssetType.REAL_ESTATE,
    "real_estate": AssetType.REAL_ESTATE,
}

# 资产状态映射
ASSET_STATUS_MAP = {
    "在库": AssetStatus.STOCK,
    "stock": AssetStatus.STOCK,
    "使用中": AssetStatus.IN_USE,
    "in_use": AssetStatus.IN_USE,
    "维修中": AssetStatus.REPAIR,
    "repair": AssetStatus.REPAIR,
    "已报废": AssetStatus.SCRAPPED,
    "scrapped": AssetStatus.SCRAPPED,
    "丢失": AssetStatus.LOST,
    "lost": AssetStatus.LOST,
}


def parse_date(value: Any) -> Optional[datetime]:
    """解析日期值"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        # Excel 日期序列号
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value)
        except Exception:
            return None
    # 字符串格式
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%d/%m/%Y"]:
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


def parse_float(value: Any) -> Optional[float]:
    """解析浮点数"""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_int(value: Any) -> Optional[int]:
    """解析整数"""
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


class ImportError:
    """导入错误"""

    def __init__(self, row: int, field: str, message: str, value: Any = None):
        self.row = row
        self.field = field
        self.message = message
        self.value = value

    def to_dict(self) -> dict:
        return {
            "row": self.row,
            "field": self.field,
            "message": self.message,
            "value": str(self.value) if self.value is not None else None,
        }


class AssetImportResult:
    """资产批量导入结果"""

    def __init__(
        self,
        total: int,
        success: int,
        failed: int,
        errors: Optional[List[ImportError]] = None,
        imported_assets: Optional[List[Dict[str, Any]]] = None,
    ):
        self.total = total
        self.success = success
        self.failed = failed
        self.errors = errors or []
        self.imported_assets = imported_assets or []

    def to_dict(self) -> dict:
        return {
            "total": self.total,
            "success": self.success,
            "failed": self.failed,
            "errors": [e.to_dict() for e in self.errors],
            "imported_assets": self.imported_assets,
        }


def validate_asset_row(
    row_data: Dict[str, Any],
    row_num: int,
    db: Session,
) -> Tuple[Optional[Dict[str, Any]], List[ImportError]]:
    """
    验证并转换单行资产数据

    Returns:
        (validated_data, errors) - 验证通过返回 (data, []), 失败返回 (None, [errors])
    """
    errors: List[ImportError] = []
    data: Dict[str, Any] = {}

    # 必填字段检查
    name = str(row_data.get("name") or "").strip()
    if not name:
        errors.append(ImportError(row_num, "name", "资产名称不能为空"))
    else:
        data["name"] = name

    # 资产类型
    asset_type_str = str(row_data.get("asset_type") or "fixed").strip().lower()
    asset_type = ASSET_TYPE_MAP.get(asset_type_str)
    if asset_type is None:
        errors.append(
            ImportError(
                row_num,
                "asset_type",
                f"无效的资产类型: {asset_type_str}，可选: 固定资产/fixed, 易耗品/consumable, 房地产/real_estate",
                asset_type_str,
            )
        )
    else:
        data["asset_type"] = asset_type

    # 分类编码 -> 分类ID
    category_code = str(row_data.get("category_code") or "").strip()
    if category_code:
        category = db.query(Category).filter(Category.code == category_code).first()
        if category:
            data["category_id"] = category.id
            # 如果 Excel 未指定资产类型，使用分类的类型
            if "asset_type" not in data:
                data["asset_type"] = category.asset_type
        else:
            errors.append(
                ImportError(row_num, "category_code", f"分类编码不存在: {category_code}")
            )

    # 部门编码 -> 部门ID
    dept_code = str(row_data.get("department_code") or "").strip()
    if dept_code:
        dept = db.query(Department).filter(Department.code == dept_code).first()
        if dept:
            data["department_id"] = dept.id
        else:
            errors.append(
                ImportError(row_num, "department_code", f"部门编码不存在: {dept_code}")
            )

    # 品牌/型号/序列号
    data["brand"] = str(row_data.get("brand") or "").strip() or None
    data["model"] = str(row_data.get("model") or "").strip() or None
    data["serial_no"] = str(row_data.get("serial_no") or "").strip() or None

    # 日期字段
    purchase_date = parse_date(row_data.get("purchase_date"))
    data["purchase_date"] = purchase_date
    warranty_end = parse_date(row_data.get("warranty_end"))
    data["warranty_end"] = warranty_end

    # 价格
    purchase_price = parse_float(row_data.get("purchase_price"))
    data["purchase_price"] = purchase_price or 0.0

    # 位置
    data["location"] = str(row_data.get("location") or "").strip() or None

    # 数量
    quantity = parse_int(row_data.get("quantity"))
    data["quantity"] = quantity if quantity is not None else 1

    # 单位
    data["unit"] = str(row_data.get("unit") or "").strip() or None

    # 描述/备注
    data["description"] = str(row_data.get("description") or "").strip() or None
    data["remarks"] = str(row_data.get("remarks") or "").strip() or None

    return (None, errors) if errors else (data, [])


def parse_excel_file(
    file_content: bytes,
    sheet_name: Optional[str] = None,
) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
    """
    解析 Excel 文件

    Returns:
        (rows, error_message) - 解析成功返回 (rows, None), 失败返回 (None, error)
    """
    from io import BytesIO
    try:
        wb = load_workbook(BytesIO(file_content), data_only=True)
    except InvalidFileException as e:
        return None, f"无效的 Excel 文件: {e}"
    except Exception as e:
        return None, f"无法读取 Excel 文件: {e}"

    # 选择工作表
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # 默认使用第一个工作表
        if not wb.sheetnames:
            return None, "Excel 文件中没有工作表"
        ws = wb[wb.sheetnames[0]]

    # 解析表头
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, "Excel 文件为空"

    header_row = rows[0]
    if not header_row or all(v is None for v in header_row):
        return None, "Excel 文件表头为空"

    # 构建 header -> index 映射（忽略大小写）
    header_map: Dict[str, int] = {}
    for idx, col_name in enumerate(header_row):
        if col_name is not None:
            header_map[str(col_name).strip()] = idx

    # 验证必需列
    required_cols = {"资产名称", "分类编码", "资产类型"}
    missing = required_cols - set(header_map.keys())
    if missing:
        return None, f"缺少必需列: {', '.join(missing)}，模板列: {', '.join(h for h, _ in ASSET_IMPORT_HEADERS)}"

    # 解析数据行
    data_rows: List[Dict[str, Any]] = []
    for row_idx, row in enumerate(rows[1:], start=2):  # 从第2行开始（第1行是表头）
        # 跳过空行
        if not row or all(v is None for v in row):
            continue

        row_data: Dict[str, Any] = {}
        for col_name, field_name in ASSET_IMPORT_HEADERS:
            if col_name in header_map:
                row_data[field_name] = row[header_map[col_name]]
        data_rows.append(row_data)

    return data_rows, None


def generate_import_template() -> bytes:
    """
    生成导入模板 Excel 文件

    Returns:
        Excel 文件字节数据
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "资产导入模板"

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写入表头
    headers = [h for h, _ in ASSET_IMPORT_HEADERS]
    required_indices = {0, 1, 2}  # 资产名称, 分类编码, 资产类型 为必填
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        # 必填列标注
        if col_idx - 1 in required_indices:
            cell.comment = None  # 可扩展添加批注

    # 设置列宽
    col_widths = [20, 15, 12, 15, 15, 20, 15, 12, 15, 15, 20, 8, 8, 20, 20]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    # 添加示例数据行
    sample_data = [
        "示例资产A", "ELECTRONIC", "固定资产", "联想", "ThinkPad X1", "SN12345678",
        "2024-01-15", "8999.00", "2027-01-15", "TECH", "A栋101", 1, "台",
        "采购自联想官网", "成色良好",
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # 添加说明sheet
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ("填写说明", ""),
        ("1. 请勿修改表头名称和顺序", ""),
        ("2. * 号列为必填项，其他列为选填", ""),
        ("3. 资产类型可选: 固定资产/fixed, 易耗品/consumable, 房地产/real_estate", ""),
        ("4. 分类编码和部门编码请确保已系统中存在", ""),
        ("5. 日期格式: YYYY-MM-DD 或 YYYY/MM/DD", ""),
        ("6. 购买价格为数字，如 8999.00", ""),
        ("7. 如有其他问题请联系管理员", ""),
    ]
    for row_idx, (key, value) in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_idx, column=1, value=key)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)

    from io import BytesIO

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
