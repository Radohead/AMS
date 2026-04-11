"""
批量导出模块测试

用例编号:
- TEST-EXP-001: test_export_assets_xlsx        - 导出资产为 Excel 格式
- TEST-EXP-002: test_export_assets_csv          - 导出资产为 CSV 格式
- TEST-EXP-003: test_export_with_filter        - 带筛选条件导出
- TEST-EXP-004: test_export_invalid_format     - 无效导出格式
- TEST-EXP-005: test_export_without_permission - 无导出权限
- TEST-EXP-006: test_export_empty_result        - 导出空结果
"""
import pytest
from io import BytesIO
from openpyxl import load_workbook


class TestBatchExport:
    """批量导出功能测试"""

    def test_export_assets_xlsx(
        self, client, admin_user, test_category, db_session
    ):
        """TEST-EXP-001: 导出资产为 Excel 格式"""
        # 登录获取 token
        login_resp = client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin123"}
        )
        token = login_resp.json()["access_token"]

        # 创建测试资产
        from app.models.asset import Asset, AssetStatus, AssetType
        from app.core.security import get_password_hash
        from app.models.user import User

        user = db_session.query(User).filter(User.username == "admin").first()
        asset = Asset(
            asset_no="TEST001",
            name="测试资产",
            category_id=test_category.id,
            asset_type=AssetType.FIXED,
            status=AssetStatus.STOCK,
            brand="联想",
            model="ThinkPad",
            purchase_price=5000.0,
            created_by=user.id if user else None,
        )
        db_session.add(asset)
        db_session.commit()

        # 导出
        response = client.post(
            "/api/assets/export?format=xlsx",
            headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats"
        )
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]

        # 验证 Excel 内容
        content = BytesIO(response.content)
        wb = load_workbook(content)
        ws = wb.active

        # 检查表头
        headers = [cell.value for cell in ws[1]]
        assert "资产编码" in headers
        assert "资产名称" in headers
        assert "分类" in headers

        # 检查数据
        assert ws.max_row >= 2  # 至少有一行数据
        assert any(row[0].value == "TEST001" for row in ws.iter_rows(min_row=2))

    def test_export_assets_csv(
        self, client, admin_user, test_category, db_session
    ):
        """TEST-EXP-002: 导出资产为 CSV 格式"""
        login_resp = client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin123"}
        )
        token = login_resp.json()["access_token"]

        # 导出 CSV
        response = client.post(
            "/api/assets/export?format=csv",
            headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert ".csv" in response.headers["content-disposition"]

        # 验证 CSV 内容
        content = response.content.decode("utf-8-sig")
        lines = content.strip().split("\n")
        assert len(lines) >= 1  # 至少有表头

        # 检查表头包含必要字段
        headers = lines[0].split(",")
        assert any("资产编码" in h for h in headers)
        assert any("资产名称" in h for h in headers)

    def test_export_with_filter(
        self, client, admin_user, test_category, db_session
    ):
        """TEST-EXP-003: 带筛选条件导出"""
        login_resp = client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin123"}
        )
        token = login_resp.json()["access_token"]

        # 按分类筛选导出
        response = client.post(
            f"/api/assets/export?format=xlsx&category_id={test_category.id}",
            headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code == 200

        # 验证只包含指定分类的资产
        content = BytesIO(response.content)
        wb = load_workbook(content)
        ws = wb.active

        # 检查分类列
        headers = [cell.value for cell in ws[1]]
        category_idx = headers.index("分类") if "分类" in headers else None
        if category_idx:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[category_idx]:  # 非空行
                    assert test_category.name in str(row[category_idx])

    def test_export_invalid_format(
        self, client, admin_user
    ):
        """TEST-EXP-004: 无效导出格式"""
        login_resp = client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin123"}
        )
        token = login_resp.json()["access_token"]

        response = client.post(
            "/api/assets/export?format=pdf",
            headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code == 400
        assert "不支持的导出格式" in response.json()["detail"]

    def test_export_without_permission(
        self, client, test_user
    ):
        """TEST-EXP-005: 无导出权限"""
        # test_user 没有资产读取权限
        login_resp = client.post(
            "/api/auth/login",
            json={"username": "testuser", "password": "testpass123"}
        )
        token = login_resp.json()["access_token"]

        response = client.post(
            "/api/assets/export?format=xlsx",
            headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code in (403, 401)

    def test_export_empty_result(
        self, client, admin_user, db_session
    ):
        """TEST-EXP-006: 导出空结果"""
        login_resp = client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin123"}
        )
        token = login_resp.json()["access_token"]

        # 导出不存在的资产
        response = client.post(
            "/api/assets/export?format=xlsx&keyword=NONEXISTENT12345",
            headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code == 200

        # 验证返回的是空数据（只有表头）
        content = BytesIO(response.content)
        wb = load_workbook(content)
        ws = wb.active

        assert ws.max_row == 1  # 只有表头，无数据行

    def test_export_all_assets(
        self, client, admin_user, test_category, db_session
    ):
        """TEST-EXP-007: 导出全部资产（无筛选条件）"""
        login_resp = client.post(
            "/api/auth/login",
            json={"username": "admin", "password": "admin123"}
        )
        token = login_resp.json()["access_token"]

        # 创建多条测试资产
        from app.models.asset import Asset, AssetStatus, AssetType
        from app.models.user import User

        user = db_session.query(User).filter(User.username == "admin").first()
        for i in range(3):
            asset = Asset(
                asset_no=f"BATCH{i:03d}",
                name=f"批量导出测试{i}",
                category_id=test_category.id,
                asset_type=AssetType.FIXED,
                status=AssetStatus.STOCK,
                purchase_price=1000.0 * (i + 1),
                created_by=user.id if user else None,
            )
            db_session.add(asset)
        db_session.commit()

        # 导出全部
        response = client.post(
            "/api/assets/export?format=xlsx",
            headers={"Authorization": f"Bearer {token}"}
        )

        assert response.status_code == 200

        # 验证包含所有创建的资产
        content = BytesIO(response.content)
        wb = load_workbook(content)
        ws = wb.active

        # 至少有 4 行（表头 + 3条测试数据）
        assert ws.max_row >= 4

        # 检查导出文件大小合理
        assert len(response.content) > 0
